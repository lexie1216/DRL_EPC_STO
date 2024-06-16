# -*- ecoding: utf-8 -*-
# @ModuleName: evaluation_t
# @Function: 
# @Author: Lexie
# @Time: 2024/5/30 16:16
import pickle
import numpy as np
import pandas as pd
from scipy import stats
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

R0 = 'high'

city = 'sz'
# city = 'tokyo'
# city = 'nyc'

# 读取字典
with open(f'res/evaluation_index_{city}_{R0}.pickle', 'rb') as f:
    d = pickle.load(f)


def my_t(sample1, sample2):
    sample1 = np.asarray(sample1)
    sample2 = np.asarray(sample2)
    r = stats.ttest_ind(sample1, sample2)
    return r.statistic, r.pvalue


# 获取所有的场景和评价指标，不包括'reward'
scenarios = list(d.keys())
metrics = list(d[scenarios[0]].keys())
if 'reward' in metrics:
    metrics.remove('reward')


# 计算每个场景每个指标的均值，保留三位小数
mean_values = {}
for scenario in scenarios:
    mean_values[scenario] = {}
    for metric in metrics:
        if metric == 'IOR':
            mean_values[scenario][metric] = round(np.mean(d[scenario][metric]) * 100, 2)  # 直接转换为百分数并保留两位小数
        else:
            mean_values[scenario][metric] = round(np.mean(d[scenario][metric]), 3)


# 计算每个场景每个指标的方差，保留10位小数
variance_values = {scenario: {metric: round(np.var(d[scenario][metric]), 10) for metric in metrics} for scenario in scenarios}

# 存储每个评价指标在不同场景下的最低场景
significant_lowest_scenarios = {metric: [] for metric in metrics}

# 比较每个评价指标在不同场景下的值
for metric in metrics:
    avg_values = {scenario: mean_values[scenario][metric] for scenario in scenarios}
    sorted_scenarios = sorted(avg_values, key=avg_values.get)

    lowest_scenarios = [sorted_scenarios[0]]

    for i in range(1, len(sorted_scenarios)):
        next_scenario = sorted_scenarios[i]

        # 对已经确定为最低的场景和下一个场景进行t检验
        all_significant = True
        for scenario in lowest_scenarios:
            stat, pvalue = my_t(d[scenario][metric], d[next_scenario][metric])

            if pvalue <= 0.001:
                all_significant = False
                break

        if all_significant:
            lowest_scenarios.append(next_scenario)
        else:
            break

    significant_lowest_scenarios[metric] = lowest_scenarios

# 创建一个DataFrame存储均值
df_mean = pd.DataFrame(mean_values).T
df_variance = pd.DataFrame(variance_values).T

# 创建Excel Writer对象
with pd.ExcelWriter(f'evaluation_results_{city}_{R0}.xlsx', engine='openpyxl') as writer:
    # 输出均值
    df_mean.to_excel(writer, sheet_name='Mean_Results')
    workbook = writer.book
    sheet = workbook['Mean_Results']

    # 设置标题样式
    title_font = Font(name='Times New Roman', size=14, bold=True)
    for cell in sheet["1:1"]:  # 第一行
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):  # 第一列
        for cell in row:
            cell.font = title_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 调整单元格样式
    normal_font = Font(name='Times New Roman', size=14)
    for col_idx, metric in enumerate(metrics, start=2):  # 从第2列开始
        best_scenarios = significant_lowest_scenarios[metric]
        for row_idx, scenario in enumerate(scenarios, start=2):  # 从第2行开始
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = normal_font
            if scenario in best_scenarios:
                cell.font = Font(name='Times New Roman', size=14, bold=True)
            # 如果是IOR列，将其格式化为百分数
            if metric == 'IOR':
                cell.value = f"{cell.value}%"

    # 调整列宽
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列字母
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    # 调整行高
    default_height = 15  # 默认行高
    for row in range(1, sheet.max_row + 1):
        current_height = sheet.row_dimensions[row].height
        if current_height is None:
            current_height = default_height
        sheet.row_dimensions[row].height = current_height * 2

    # 输出方差
    df_variance.to_excel(writer, sheet_name='Variance_Results')
    sheet_var = workbook['Variance_Results']

    # 设置标题样式
    for cell in sheet_var["1:1"]:  # 第一行
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in sheet_var.iter_rows(min_row=2, max_row=sheet_var.max_row, min_col=1, max_col=1):  # 第一列
        for cell in row:
            cell.font = title_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 调整单元格样式
    for col_idx, metric in enumerate(metrics, start=2):  # 从第2列开始
        for row_idx, scenario in enumerate(scenarios, start=2):  # 从第2行开始
            cell = sheet_var.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = normal_font
            # 如果是IOR列，将其格式化为百分数
            if metric == 'IOR':
                cell.value = f"{cell.value}%"

    # 调整列宽
    for col in sheet_var.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列字母
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet_var.column_dimensions[column].width = adjusted_width

    # 调整行高
    for row in range(1, sheet_var.max_row + 1):
        current_height = sheet_var.row_dimensions[row].height
        if current_height is None:
            current_height = default_height
        sheet_var.row_dimensions[row].height = current_height * 2

print(f'Excel表格已生成并保存为evaluation_results_{city}_{R0}.xlsx')
